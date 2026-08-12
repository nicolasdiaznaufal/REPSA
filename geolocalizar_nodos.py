"""
Construye la tabla `nodos_ubicacion` en pml.duckdb, cruzando el catálogo
de nodos de CENACE con los centroides de municipios de INEGI.

Uso:
    python geolocalizar_nodos.py "Catálogo NodosP Sistema Eléctrico Nacional (v...).xlsx"

Fuentes:
- Catálogo de nodos: https://www.cenace.gob.mx/Paginas/SIM/NodosP.aspx
  (descarga la versión más reciente y pásala como argumento)
- Centroides de municipios (cabecera municipal, INEGI): se incluyen ya
  procesados en municipios_centroides.csv en este mismo folder. Si CENACE
  actualiza los límites de municipios, ese archivo rara vez necesita cambiar.

Nota sobre precisión: la ubicación es a nivel de MUNICIPIO (la cabecera
municipal), no la coordenada exacta de la subestación. Es la granularidad
más fina que CENACE publica en su catálogo público.
"""
import sys
import os
import openpyxl
import pandas as pd
import duckdb

DB_PATH = os.path.join(os.path.dirname(__file__), ".", "db", "pml.duckdb")
CENTROIDES_PATH = os.path.join(os.path.dirname(__file__), "municipios_centroides.csv")


def cargar_catalogo(path_xlsx):
    wb = openpyxl.load_workbook(path_xlsx, data_only=True)
    # El nombre de la hoja cambia con cada versión (incluye la fecha), tomamos la primera
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(min_row=3, values_only=True))
    cols = ['sistema', 'centro_control_regional', 'zona_carga', 'clave', 'nombre',
            'nivel_tension_kv', 'carga_directa', 'carga_indirecta', 'gen_directa',
            'gen_indirecta', 'zona_operacion_transmision', 'gerencia_regional_transmision',
            'zona_distribucion', 'gerencia_divisional_distribucion', 'clave_entidad',
            'entidad_federativa', 'clave_municipio', 'municipio', 'region_transmision']
    cat = pd.DataFrame(rows, columns=cols)

    cat['clave_entidad_num'] = pd.to_numeric(cat['clave_entidad'], errors='coerce')
    cat['clave_municipio_num'] = pd.to_numeric(cat['clave_municipio'], errors='coerce')
    cat['clave_entidad_str'] = cat['clave_entidad_num'].apply(
        lambda x: str(int(x)).zfill(2) if pd.notna(x) else None)
    cat['clave_municipio_str'] = cat['clave_municipio_num'].apply(
        lambda x: str(int(x)).zfill(3) if pd.notna(x) else None)
    return cat


def geolocalizar(cat):
    centroides = pd.read_csv(CENTROIDES_PATH, dtype={'CVE_ENT': str, 'CVE_MUN': str})
    nodos_geo = cat.merge(
        centroides,
        left_on=['clave_entidad_str', 'clave_municipio_str'],
        right_on=['CVE_ENT', 'CVE_MUN'],
        how='left'
    )
    return nodos_geo


def guardar(nodos_geo):
    geo_limpio = nodos_geo[[
        'clave', 'nombre', 'centro_control_regional', 'entidad_federativa', 'municipio',
        'nivel_tension_kv', 'LAT_DEC', 'LON_DEC'
    ]].rename(columns={'clave': 'nodo', 'LAT_DEC': 'lat', 'LON_DEC': 'lon'})
    geo_limpio = geo_limpio.dropna(subset=['lat', 'lon'])

    con = duckdb.connect(DB_PATH)
    con.register('geo_df', geo_limpio)
    con.execute('CREATE OR REPLACE TABLE nodos_ubicacion AS SELECT * FROM geo_df')
    con.execute('CREATE UNIQUE INDEX IF NOT EXISTS idx_nodo_ubicacion ON nodos_ubicacion(nodo)')
    total = con.execute('SELECT COUNT(*) FROM nodos_ubicacion').fetchone()[0]
    con.close()
    return total, len(geo_limpio)


def main(path_xlsx):
    print(f"Cargando catálogo: {path_xlsx}")
    cat = cargar_catalogo(path_xlsx)
    print(f"  -> {len(cat)} nodos en el catálogo")

    print("Geolocalizando por municipio...")
    nodos_geo = geolocalizar(cat)
    sin_match = nodos_geo['LAT_DEC'].isna().sum()
    print(f"  -> {len(nodos_geo) - sin_match} de {len(nodos_geo)} nodos geolocalizados "
          f"({sin_match} sin coincidencia)")

    total, guardados = guardar(nodos_geo)
    print(f"\nTabla 'nodos_ubicacion' actualizada en {DB_PATH}: {guardados} nodos con coordenadas.")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print(__doc__)
        sys.exit(1)
    main(sys.argv[1])
    